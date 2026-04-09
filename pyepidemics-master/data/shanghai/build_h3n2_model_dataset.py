from __future__ import annotations

from datetime import date
from statistics import mean

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from build_shanghai_h3n2_proxy import make_daily_rows


OUTPUT_PATH = r"D:\HP\Downloads\上海H3N2_SEIR代理数据集_2017.xlsx"

POPULATION = 24_183_300
TOTAL_ATTACK_RATE = 0.03
TOTAL_INFECTIONS = round(POPULATION * TOTAL_ATTACK_RATE)
EFFECTIVE_SUSCEPTIBLE_SHARE = 0.25
EFFECTIVE_S0 = round(POPULATION * EFFECTIVE_SUSCEPTIBLE_SHARE)
ASYMPTOMATIC_SHARE = 0.16
SYMPTOMATIC_SHARE = 1.0 - ASYMPTOMATIC_SHARE
CARE_SEEKING_SHARE = 0.25
LATENT_DAYS = 1.5
INFECTIOUS_DAYS = 3.0
SIGMA = 1.0 / LATENT_DAYS
GAMMA = 1.0 / INFECTIOUS_DAYS

URL_POP = "https://www.shanghai.gov.cn/nw9822/20200906/0001-9822_1283583.html"
URL_WHO_ATTACK = "https://www.who.int/publications/m/item/vaccine-preventable-diseases-surveillance-standards-influenza"
URL_WHO_INCUBATION = "https://www.who.int/news-room/fact-sheets/detail/influenza-%28seasonal%EF%BC%89"
URL_SERIAL = "https://pmc.ncbi.nlm.nih.gov/articles/PMC3057478/"
URL_ASYM = "https://pmc.ncbi.nlm.nih.gov/articles/PMC4586318/"
URL_CARE = "https://pmc.ncbi.nlm.nih.gov/articles/PMC9850268/"
URL_PROXY_PAPER_1 = "https://pmc.ncbi.nlm.nih.gov/articles/PMC11378670/"
URL_PROXY_PAPER_2 = "https://www.nature.com/articles/s41598-022-19228-y"
URL_W30 = "https://ivdc.chinacdc.cn/cnic/zyzx/lgzb/201708/t20170807_149211.htm"
URL_W33 = "https://ivdc.chinacdc.cn/cnic/zyzx/lgzb/201708/t20170828_151328.htm"
URL_W34 = "https://ivdc.chinacdc.cn/cnic/zyzx/lgzb/201709/t20170905_151982.htm"
URL_W38 = "https://ivdc.chinacdc.cn/cnic/zyzx/lgzb/201711/t20171106_154688.htm"


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="0F4C5C")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def fit_widths(ws, max_width: int = 44) -> None:
    for idx, col in enumerate(ws.columns, start=1):
        values = ["" if c.value is None else str(c.value) for c in col]
        width = min(max(len(v) for v in values) + 2, max_width)
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_daily_model_rows() -> list[dict[str, object]]:
    proxy_rows = make_daily_rows()
    total_weight = sum(float(r["上海H3N2代理阳性率(%)"]) for r in proxy_rows)

    infections = []
    allocated = 0
    for idx, row in enumerate(proxy_rows):
        weight = float(row["上海H3N2代理阳性率(%)"]) / total_weight if total_weight else 0.0
        if idx < len(proxy_rows) - 1:
            inf = round(TOTAL_INFECTIONS * weight)
            allocated += inf
        else:
            inf = TOTAL_INFECTIONS - allocated
        infections.append(inf)

    rows: list[dict[str, object]] = []
    s_eff = float(EFFECTIVE_S0)
    e = 0.0
    i = 0.0
    r = 0.0
    cum_inf = 0
    cum_rep = 0

    for proxy, new_inf in zip(proxy_rows, infections):
        if i > 0 and s_eff > 0:
            beta_eff = new_inf * EFFECTIVE_S0 / (s_eff * i)
            rt_eff = beta_eff * s_eff / (GAMMA * EFFECTIVE_S0)
        else:
            beta_eff = None
            rt_eff = None

        new_sym = round(new_inf * SYMPTOMATIC_SHARE)
        new_rep = round(new_sym * CARE_SEEKING_SHARE)
        new_asym = new_inf - new_sym

        to_i = e * SIGMA
        to_r = i * GAMMA
        e = max(e + new_inf - to_i, 0.0)
        i = max(i + to_i - to_r, 0.0)
        r = r + to_r
        s_eff = max(s_eff - new_inf, 0.0)

        cum_inf += new_inf
        cum_rep += new_rep

        observed_case_rate_100k = new_rep / POPULATION * 100000
        infection_rate_100k = new_inf / POPULATION * 100000

        rows.append(
            {
                "日期": proxy["日期"],
                "年份": proxy["年份"],
                "月份": proxy["月份"],
                "ISO周": proxy["ISO周"],
                "上海H3N2代理阳性率(%)": proxy["上海H3N2代理阳性率(%)"],
                "新感染数": new_inf,
                "新有症状感染数": new_sym,
                "新无症状感染数": new_asym,
                "新报告病例数": new_rep,
                "累计感染数": cum_inf,
                "累计报告病例数": cum_rep,
                "有效易感数_S": round(s_eff),
                "潜伏者_E": round(e),
                "传染者_I": round(i),
                "移除者_R": round(r),
                "日感染率_每10万人": round(infection_rate_100k, 2),
                "日报告发病率_每10万人": round(observed_case_rate_100k, 2),
                "beta_eff": None if beta_eff is None else round(beta_eff, 4),
                "Rt_eff": None if rt_eff is None else round(rt_eff, 3),
                "证据等级": proxy["证据等级"],
                "插值区间": proxy["插值区间"],
                "阶段": proxy["阶段"],
            }
        )
    return rows


def add_overview(ws, rows: list[dict[str, object]]) -> None:
    peak_inf = max(rows, key=lambda r: r["新感染数"])
    peak_i = max(rows, key=lambda r: r["传染者_I"])
    avg_rt = mean(r["Rt_eff"] for r in rows if r["Rt_eff"] is not None)

    ws["A1"] = "上海 H3N2 2017 SEIR 代理数据集"
    ws["A1"].font = Font(size=16, bold=True)
    entries = [
        ("数据性质", "基于公开周报与论文重建的模型输入数据，不是官方原始病例台账"),
        ("总人口 N", POPULATION),
        ("假设波次总感染率", TOTAL_ATTACK_RATE),
        ("假设波次总感染数", TOTAL_INFECTIONS),
        ("有效易感池初值 S0", EFFECTIVE_S0),
        ("峰值新感染日期", peak_inf["日期"]),
        ("峰值新感染数", peak_inf["新感染数"]),
        ("峰值传染者日期", peak_i["日期"]),
        ("峰值传染者 I", peak_i["传染者_I"]),
        ("平均 Rt_eff", round(avg_rt, 3)),
    ]
    fill = PatternFill("solid", fgColor="DDEBF7")
    for row_idx, (k, v) in enumerate(entries, start=3):
        ws[f"A{row_idx}"] = k
        ws[f"B{row_idx}"] = v
        ws[f"A{row_idx}"].font = Font(bold=True)
        ws[f"A{row_idx}"].fill = fill
        ws[f"B{row_idx}"].alignment = Alignment(wrap_text=True)
    ws["B5"].number_format = "0.0%"

    notes = [
        "1. 当前文件把代理阳性率映射成 SEIR 状态量，供动力学建模使用。",
        "2. 感染总量、有效易感池、就诊/报告比例属于模型假设，不是直接观测值。",
        "3. 若要用于参数反演，建议把总感染率和有效易感池作为待校准参数做敏感性分析。",
        "4. 没有足够公开资料支持严谨重建区县分层或年龄分层病例数，因此本版只给总量口径。",
    ]
    ws["A15"] = "使用边界"
    ws["A15"].font = Font(bold=True)
    for idx, note in enumerate(notes, start=16):
        ws[f"A{idx}"] = note

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 46


def add_parameter_sheet(ws) -> None:
    ws.append(["参数", "取值", "用途", "依据", "URL"])
    rows = [
        ["总人口 N", POPULATION, "城市总人口基数", "上海市 2017 年末常住人口 2418.33 万", URL_POP],
        ["波次总感染率", TOTAL_ATTACK_RATE, "把比例曲线映射为总感染量", "假设值；低于 WHO 对完整季节流感的成年人 5%-10% 年攻击率，因为这里只取单个 H3N2 夏季波次", URL_WHO_ATTACK],
        ["有效易感池占比", EFFECTIVE_SUSCEPTIBLE_SHARE, "使季节性免疫背景下的 S 具有可解释性", "假设值；用于表示进入本次 H3N2 夏季波次的有效易感人口，而非全人口均完全易感", URL_PROXY_PAPER_1],
        ["无症状比例", ASYMPTOMATIC_SHARE, "感染者拆分为有症状与无症状", "Leung 等系统综述：PCR/暴发调查研究的无症状比例 pooled mean 约 16%", URL_ASYM],
        ["有症状比例", SYMPTOMATIC_SHARE, "估算有症状感染数", "由 1 - 无症状比例得到", URL_ASYM],
        ["就诊/报告比例", CARE_SEEKING_SHARE, "由有症状感染映射为报告病例", "ILI 就诊研究显示约 24.1%-31.1% 寻求医疗；此处取 25% 作为基线假设", URL_CARE],
        ["潜伏期", LATENT_DAYS, "SEIR 中 E->I 速率", "WHO 说明流感感染到发病约 2 天；模型中取 1.5 天潜伏期，使生成时间与文献估计接近", URL_WHO_INCUBATION],
        ["传染期", INFECTIOUS_DAYS, "SEIR 中 I->R 速率", "结合平均序列间隔约 3.6 天的研究，取 3 天传染期作为简化假设", URL_SERIAL],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    fit_widths(ws, max_width=68)
    for cell in ws["B"][1:]:
        if isinstance(cell.value, float) and cell.value < 1:
            cell.number_format = "0.0%"


def add_daily_sheet(ws, rows: list[dict[str, object]]) -> None:
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])
    style_sheet(ws)
    fit_widths(ws)

    for cell in ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for col in ["E"]:
        for cell in ws[col][1:]:
            cell.number_format = "0.00"
    for col in ["P", "Q"]:
        for cell in ws[col][1:]:
            cell.number_format = "0.00"


def add_method_sheet(ws) -> None:
    ws.append(["步骤", "说明"])
    rows = [
        ["1. 代理曲线", "使用上一版工作簿中的上海 H3N2 代理阳性率，曲线锚定于国家流感中心周报和上海论文。"],
        ["2. 分配总感染数", "按每日代理阳性率在整个波次中的占比，分配假设的总感染数 725,499。"],
        ["3. 症状拆分", "新有症状感染数 = 新感染数 × 84%；新无症状感染数 = 新感染数 × 16%。"],
        ["4. 报告病例映射", "新报告病例数 = 新有症状感染数 × 25%；表示被医疗系统捕获的病例口径。"],
        ["5. SEIR 递推", "E_t = E_(t-1) + new_inf_t - sigma*E_(t-1)；I_t = I_(t-1) + sigma*E_(t-1) - gamma*I_(t-1)。"],
        ["6. 有效易感池", "S_t 只在有效易感池内递减；这比用全市总人口直接作为完全易感更符合季节性流感背景。"],
        ["7. beta_eff 与 Rt_eff", "由 incidence = beta_eff * S * I / S0_eff 反推，得到时间变化的有效传播强度，仅供校准初值使用。"],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    fit_widths(ws, max_width=72)


def add_sources_sheet(ws) -> None:
    ws.append(["来源类别", "出处", "关键用途", "URL"])
    rows = [
        ["人口", "上海市 2017 年国民经济运行情况", "总人口 N", URL_POP],
        ["官方周报", "国家流感中心 2017 第30周", "代理曲线锚点", URL_W30],
        ["官方周报", "国家流感中心 2017 第33周", "代理曲线锚点", URL_W33],
        ["官方周报", "国家流感中心 2017 第34周", "代理曲线锚点", URL_W34],
        ["官方周报", "国家流感中心 2017 第38周", "代理曲线锚点", URL_W38],
        ["上海论文", "Emerg Microbes Infect. 2024", "2017 上海夏季峰时间约束", URL_PROXY_PAPER_1],
        ["上海论文", "Scientific Reports 2022", "2017 上海口岸 H3N2 夏季波次约束", URL_PROXY_PAPER_2],
        ["WHO", "Influenza surveillance standards", "完整季节攻击率参考区间", URL_WHO_ATTACK],
        ["WHO", "Seasonal influenza fact sheet", "感染到发病 1-4 天", URL_WHO_INCUBATION],
        ["综述", "Influenza asymptomatic fraction meta-analysis", "无症状比例", URL_ASYM],
        ["实证研究", "ILI care-seeking study", "就诊/报告比例基线", URL_CARE],
        ["实证研究", "Serial interval of influenza", "序列间隔与传染期简化", URL_SERIAL],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    fit_widths(ws, max_width=72)


def add_chart(ws_daily, ws_chart, n_rows: int) -> None:
    chart1 = LineChart()
    chart1.title = "新感染数与新报告病例数"
    chart1.y_axis.title = "人数"
    chart1.x_axis.title = "日期"
    data1 = Reference(ws_daily, min_col=6, max_col=9, min_row=1, max_row=n_rows + 1)
    cats = Reference(ws_daily, min_col=1, min_row=2, max_row=n_rows + 1)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.height = 10
    chart1.width = 22
    ws_chart.add_chart(chart1, "A1")

    chart2 = LineChart()
    chart2.title = "SEIR 状态量"
    chart2.y_axis.title = "人数"
    chart2.x_axis.title = "日期"
    data2 = Reference(ws_daily, min_col=12, max_col=15, min_row=1, max_row=n_rows + 1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.height = 10
    chart2.width = 22
    ws_chart.add_chart(chart2, "A22")


def main() -> None:
    rows = build_daily_model_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "概览"
    add_overview(ws, rows)

    param = wb.create_sheet("参数与依据")
    add_parameter_sheet(param)

    daily = wb.create_sheet("每日SEIR代理")
    add_daily_sheet(daily, rows)

    method = wb.create_sheet("构建方法")
    add_method_sheet(method)

    source = wb.create_sheet("来源")
    add_sources_sheet(source)

    chart = wb.create_sheet("图表")
    add_chart(daily, chart, len(rows))

    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
