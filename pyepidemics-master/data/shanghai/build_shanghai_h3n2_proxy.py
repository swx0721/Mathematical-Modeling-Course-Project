from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from statistics import mean

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class WeeklyAnchor:
    iso_week: int
    monday: date
    south_positive_pct: float
    h3_share_pct: float
    evidence_tier: str
    note: str
    source_url: str

    @property
    def h3_proxy_pct(self) -> float:
        return round(self.south_positive_pct * self.h3_share_pct / 100.0, 2)


OUTPUT_PATH = r"D:\HP\Downloads\上海H3N2流感代理数据集_2017_更可信版.xlsx"
START_DATE = date(2017, 5, 15)
END_DATE = date(2017, 9, 15)

PAPER_2024 = "https://pmc.ncbi.nlm.nih.gov/articles/PMC11378670/"
PAPER_2022 = "https://www.nature.com/articles/s41598-022-19228-y"
W30_URL = "https://ivdc.chinacdc.cn/cnic/zyzx/lgzb/201708/t20170807_149211.htm"
W33_URL = "https://ivdc.chinacdc.cn/cnic/zyzx/lgzb/201708/t20170828_151328.htm"
W34_URL = "https://ivdc.chinacdc.cn/cnic/zyzx/lgzb/201709/t20170905_151982.htm"
W38_URL = "https://ivdc.chinacdc.cn/cnic/zyzx/lgzb/201711/t20171106_154688.htm"

ANCHORS = [
    WeeklyAnchor(
        iso_week=20,
        monday=date(2017, 5, 15),
        south_positive_pct=2.8,
        h3_share_pct=65.0,
        evidence_tier="C",
        note=(
            "窗口起点假设值。仅用于给 2017 年夏季波次设定低位起步，"
            "与上海论文所述 7-10 月夏季峰、6-8 月港口 H3N2 夏季波次保持一致。"
        ),
        source_url=PAPER_2024,
    ),
    WeeklyAnchor(
        iso_week=24,
        monday=date(2017, 6, 12),
        south_positive_pct=6.0,
        h3_share_pct=75.0,
        evidence_tier="C",
        note=(
            "6 月中旬爬升假设值。公开文献支持 2017 年夏季波次已开始形成，"
            "但无上海逐周公开数值。"
        ),
        source_url=PAPER_2022,
    ),
    WeeklyAnchor(
        iso_week=26,
        monday=date(2017, 6, 26),
        south_positive_pct=10.5,
        h3_share_pct=82.0,
        evidence_tier="C",
        note=(
            "6 月下旬加速爬升假设值，用于连接 6 月中旬与 7 月底国家流感中心已观测到的高峰状态。"
        ),
        source_url=PAPER_2022,
    ),
    WeeklyAnchor(
        iso_week=30,
        monday=date(2017, 7, 24),
        south_positive_pct=22.5,
        h3_share_pct=87.8,
        evidence_tier="A",
        note=(
            "国家流感中心 2017 第30周：南方省份流感检测阳性率 22.5%，"
            "H3N2 占南方阳性标本 574/654。"
        ),
        source_url=W30_URL,
    ),
    WeeklyAnchor(
        iso_week=33,
        monday=date(2017, 8, 14),
        south_positive_pct=25.5,
        h3_share_pct=94.4,
        evidence_tier="A",
        note=(
            "国家流感中心 2017 第33周：南方省份流感检测阳性率 25.5%，"
            "H3N2 占南方阳性标本 757/802。"
        ),
        source_url=W33_URL,
    ),
    WeeklyAnchor(
        iso_week=34,
        monday=date(2017, 8, 21),
        south_positive_pct=26.2,
        h3_share_pct=92.7,
        evidence_tier="A",
        note=(
            "国家流感中心 2017 第34周：南方省份流感检测阳性率 26.2%，"
            "H3N2 占南方阳性标本 787/849。"
        ),
        source_url=W34_URL,
    ),
    WeeklyAnchor(
        iso_week=38,
        monday=date(2017, 9, 18),
        south_positive_pct=18.0,
        h3_share_pct=91.0,
        evidence_tier="A",
        note=(
            "国家流感中心 2017 第38周：南方省份流感检测阳性率 18.0%，"
            "H3N2 占南方阳性标本 538/591。该点用于约束 9 月中旬的回落斜率。"
        ),
        source_url=W38_URL,
    ),
]


def lerp(x0: float, y0: float, x1: float, y1: float, x: float) -> float:
    if x1 == x0:
        return y0
    return y0 + (y1 - y0) * ((x - x0) / (x1 - x0))


def interpolate_for_day(current: date) -> tuple[float, float, str, str]:
    if current <= ANCHORS[0].monday:
        a = ANCHORS[0]
        return a.south_positive_pct, a.h3_share_pct, a.evidence_tier, f"W{a.iso_week}"

    if current >= ANCHORS[-1].monday:
        a = ANCHORS[-1]
        return a.south_positive_pct, a.h3_share_pct, a.evidence_tier, f"W{a.iso_week}"

    for left, right in zip(ANCHORS, ANCHORS[1:]):
        if left.monday <= current <= right.monday:
            x0 = left.monday.toordinal()
            x1 = right.monday.toordinal()
            x = current.toordinal()
            south = round(lerp(x0, left.south_positive_pct, x1, right.south_positive_pct, x), 2)
            h3 = round(lerp(x0, left.h3_share_pct, x1, right.h3_share_pct, x), 2)
            if left.evidence_tier == "A" and right.evidence_tier == "A":
                tier = "B"
            else:
                tier = "C"
            return south, h3, tier, f"W{left.iso_week}-W{right.iso_week}"
    raise RuntimeError("Interpolation failed")


def make_daily_rows() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    current = START_DATE
    while current <= END_DATE:
        south, h3_share, tier, segment = interpolate_for_day(current)
        h3_proxy = round(south * h3_share / 100.0, 2)
        rows.append(
            {
                "日期": current,
                "年份": current.year,
                "月份": current.month,
                "ISO周": current.isocalendar().week,
                "星期": current.isoweekday(),
                "是否周末": "是" if current.isoweekday() >= 6 else "否",
                "南方流感阳性率代理(%)": south,
                "H3N2占流感阳性比代理(%)": h3_share,
                "上海H3N2代理阳性率(%)": h3_proxy,
                "每100份ILI样本对应H3N2阳性当量": h3_proxy,
                "证据等级": tier,
                "插值区间": segment,
            }
        )
        current += timedelta(days=1)

    peak = max(row["上海H3N2代理阳性率(%)"] for row in rows)
    for row in rows:
        rate = float(row["上海H3N2代理阳性率(%)"])
        idx = round(rate / peak * 100.0, 1) if peak else 0.0
        row["H3N2活动指数(峰值=100)"] = idx
        if idx < 20:
            phase = "低位起步"
        elif idx < 45:
            phase = "爬升期"
        elif idx < 80:
            phase = "高位平台"
        elif idx < 95:
            phase = "峰顶附近"
        else:
            phase = "峰值区"
        row["阶段"] = phase
    return rows


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def fit_widths(ws, max_width: int = 42) -> None:
    for idx, col in enumerate(ws.columns, start=1):
        values = ["" if c.value is None else str(c.value) for c in col]
        width = min(max(len(v) for v in values) + 2, max_width)
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_overview(ws, daily_rows: list[dict[str, object]]) -> None:
    peak_row = max(daily_rows, key=lambda r: r["上海H3N2代理阳性率(%)"])
    avg_rate = round(mean(float(r["上海H3N2代理阳性率(%)"]) for r in daily_rows), 2)
    cum_equivalent = round(sum(float(r["每100份ILI样本对应H3N2阳性当量"]) for r in daily_rows), 2)

    ws["A1"] = "上海 2017 H3N2 夏季波次代理数据集"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "数据性质"
    ws["B3"] = "代理数据，不是官方原始逐日病例数据"
    ws["A4"] = "时间范围"
    ws["B4"] = f"{START_DATE.isoformat()} 至 {END_DATE.isoformat()}"
    ws["A5"] = "重建口径"
    ws["B5"] = "以上海夏季波次时间特征 + 国家流感中心南方省份周报数值锚点构建"
    ws["A6"] = "峰值日期"
    ws["B6"] = peak_row["日期"]
    ws["A7"] = "峰值代理阳性率"
    ws["B7"] = peak_row["上海H3N2代理阳性率(%)"]
    ws["A8"] = "平均代理阳性率"
    ws["B8"] = avg_rate
    ws["A9"] = "累计阳性当量"
    ws["B9"] = cum_equivalent
    ws["C9"] = "按每100份 ILI 样本累计，不等于真实病例数"

    ws["A11"] = "使用说明"
    notes = [
        "1. 每日值由逐周锚点线性插值得到，保留趋势，不制造伪精度的病例口径。",
        "2. 定量锚点优先使用国家流感中心 2017 周报中的南方省份阳性率和 H3N2 占比。",
        "3. 起始阶段缺少上海公开逐周数值，使用论文支持的季节时序进行保守设定。",
        "4. 如需病例数版本，必须先额外确定映射口径，例如哨点采样量、就诊量或住院率。 ",
    ]
    for i, note in enumerate(notes, start=12):
        ws[f"A{i}"] = note

    fill = PatternFill("solid", fgColor="D9EAF7")
    for row in range(3, 10):
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = fill
        ws[f"B{row}"].alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 32


def add_anchor_sheet(ws) -> None:
    ws.append(
        [
            "ISO周",
            "周一起始日",
            "南方流感阳性率(%)",
            "H3N2占流感阳性比(%)",
            "H3N2代理阳性率(%)",
            "证据等级",
            "说明",
            "来源URL",
        ]
    )
    for anchor in ANCHORS:
        ws.append(
            [
                anchor.iso_week,
                anchor.monday,
                anchor.south_positive_pct,
                anchor.h3_share_pct,
                anchor.h3_proxy_pct,
                anchor.evidence_tier,
                anchor.note,
                anchor.source_url,
            ]
        )
    style_sheet(ws)
    fit_widths(ws)


def add_daily_sheet(ws, daily_rows: list[dict[str, object]]) -> None:
    headers = list(daily_rows[0].keys())
    ws.append(headers)
    for row in daily_rows:
        ws.append([row[h] for h in headers])
    style_sheet(ws)
    fit_widths(ws)

    pct_cols = {
        "G": "0.0",
        "H": "0.0",
        "I": "0.0",
        "J": "0.0",
        "M": "0.0",
    }
    for col, fmt in pct_cols.items():
        for cell in ws[col][1:]:
            cell.number_format = fmt
    for cell in ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd"


def add_method_sheet(ws) -> None:
    ws.append(["字段", "内容"])
    rows = [
        ["代理目标", "重建 2017 年上海 H3N2 夏季波次的更可信日度趋势代理，不伪装成真实病例台账。"],
        ["定量锚点", "国家流感中心 2017 第30、33、34、38周周报中的南方省份流感阳性率与 H3N2 占比。"],
        ["上海约束", "2024 上海监测网络论文指出 2017 年存在 7-10 月夏季峰；2022 论文指出 2017 年 6-8 月存在 H3N2 夏季波次。"],
        ["日度展开方法", "对周一起始锚点做线性插值，不叠加工作日/周末就诊噪声。"],
        ["证据等级A", "直接来自公开周报的周度数值锚点。"],
        ["证据等级B", "两个 A 级周锚点之间的线性插值日值。"],
        ["证据等级C", "缺少公开逐周数值时，基于论文时序约束给出的保守假设。"],
        ["不提供的字段", "真实逐日确诊数、住院数、死亡数、区县分布、年龄分布。公开资料不足以严谨重建。"],
        ["推荐用途", "时间序列方法验证、波次识别、与气象或行为变量做相关分析。"],
        ["不推荐用途", "作为官方病例事实、区县排名、病死率估计、卫生资源精确负荷测算。"],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    fit_widths(ws, max_width=64)


def add_sources_sheet(ws) -> None:
    ws.append(["来源类别", "出处", "关键信息", "URL"])
    sources = [
        [
            "官方周报",
            "中国国家流感中心 2017 第30周",
            "南方阳性率 22.5%，H3N2 占南方阳性 574/654。",
            W30_URL,
        ],
        [
            "官方周报",
            "中国国家流感中心 2017 第33周",
            "南方阳性率 25.5%，H3N2 占南方阳性 757/802。",
            W33_URL,
        ],
        [
            "官方周报",
            "中国国家流感中心 2017 第34周",
            "南方阳性率 26.2%，H3N2 占南方阳性 787/849。",
            W34_URL,
        ],
        [
            "官方周报",
            "中国国家流感中心 2017 第38周",
            "南方阳性率 18.0%，H3N2 占南方阳性 538/591。",
            W38_URL,
        ],
        [
            "上海论文",
            "Emerg Microbes Infect. 2024",
            "上海 2005-2023 监测网络显示 2017 年存在 7-10 月夏季峰，流感在部分峰中可贡献 60-70% 的 ILI。",
            PAPER_2024,
        ],
        [
            "上海论文",
            "Scientific Reports 2022",
            "上海口岸 2016/2017 监测中，2017 年 6-8 月存在 H3N2 夏季波次。",
            PAPER_2022,
        ],
    ]
    for row in sources:
        ws.append(row)
    style_sheet(ws)
    fit_widths(ws, max_width=72)


def add_chart(ws_daily, ws_chart, n_rows: int) -> None:
    chart = LineChart()
    chart.title = "上海 H3N2 代理阳性率日度曲线"
    chart.y_axis.title = "代理阳性率(%)"
    chart.x_axis.title = "日期"
    data = Reference(ws_daily, min_col=9, min_row=1, max_row=n_rows + 1)
    cats = Reference(ws_daily, min_col=1, min_row=2, max_row=n_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 22
    ws_chart.add_chart(chart, "A1")
    ws_chart["A20"] = "说明：图线展示每日 H3N2 代理阳性率，不代表真实逐日确诊数。"


def main() -> None:
    daily_rows = make_daily_rows()

    wb = Workbook()
    overview = wb.active
    overview.title = "概览"
    add_overview(overview, daily_rows)

    anchors = wb.create_sheet("每周锚点")
    add_anchor_sheet(anchors)

    daily = wb.create_sheet("每日代理数据")
    add_daily_sheet(daily, daily_rows)

    method = wb.create_sheet("构建方法")
    add_method_sheet(method)

    sources = wb.create_sheet("来源")
    add_sources_sheet(sources)

    chart_ws = wb.create_sheet("图表")
    add_chart(daily, chart_ws, len(daily_rows))

    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
